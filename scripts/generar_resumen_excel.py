import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ENCABEZADOS = [
    "Fecha de Archivo", "Cuenta", "CUIT-CUIL", "Nombre-Razón Social",
    "Tipo de Documento", "Subcarpeta", "Ruta Relativa", "Hash", "Estado"
]

def generar_resumen(datos, archivo_excel, ruta_base_alternativa=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Archivos"

    # Agregar encabezados
    ws.append(ENCABEZADOS)
    for col in range(1, len(ENCABEZADOS) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Agregar filas con hipervínculos
    for fila in datos:
        ruta_relativa = fila.get("ruta_relativa", "")
        ruta_completa = (
            ruta_base_alternativa.rstrip("/") + "/" + ruta_relativa.replace("\\", "/")
            if ruta_base_alternativa else ruta_relativa
        )

        estado = (
            "Archivado" if fila.get("destino") and not fila.get("error")
            else "Duplicado" if fila.get("error") == "Duplicado por hash"
            else "Error"
        )

        nueva_fila = [
            fila.get("fecha_archivo", ""),
            fila.get("cuenta", ""),
            fila.get("cuit_cuil", ""),
            fila.get("nombre", ""),
            fila.get("tipo_documento", ""),
            fila.get("subcarpeta", ""),
            ruta_relativa,
            fila.get("hash", ""),
            estado
        ]
        ws.append(nueva_fila)

        # Insertar hipervínculo en la celda de ruta
        celda = ws.cell(row=ws.max_row, column=7)
        celda.hyperlink = ruta_completa
        celda.font = Font(color="0000FF", underline="single")

    # Ajustar ancho de columnas
    for col in range(1, len(ENCABEZADOS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(archivo_excel)
    print(f"📊 Resumen Excel generado con hash y estado en: {archivo_excel}")