import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os

ENCABEZADOS = [
    "Fecha", "Cuenta", "Nombre / Razón Social", "CUIT / CUIL",
    "Tipo de Documento", "Archivo", "Destino", "Hash"
]

def actualizar_resumen(resumen_data, excel_path="historial_archivo.xlsx", proteger=False):
    # Crear archivo si no existe
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(ENCABEZADOS)

        # Estilo de encabezados
        for col in range(1, len(ENCABEZADOS) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        # Ajuste de ancho
        for col in range(1, len(ENCABEZADOS) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    # Agregar registros nuevos
    for fila in resumen_data:
        nueva_fila = [
            fila.get("Fecha_archivo", ""),
            fila.get("Cuenta", ""),
            fila.get("Nombre-Razón Social", ""),
            fila.get("CUIT-CUIL", ""),
            fila.get("Tipo de documento", ""),
            fila.get("Subcarpeta", ""),
            fila.get("Ruta relativa", ""),
            fila.get("hash", "")
        ]
        ws.append(nueva_fila)

    # Protección opcional
    if proteger:
        ws.protection.sheet = True
        ws.protection.enable()

    wb.save(excel_path)